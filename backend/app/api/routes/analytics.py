from __future__ import annotations

from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response, StreamingResponse
from motor.motor_asyncio import AsyncIOMotorDatabase

from app.core.dependencies import get_db, require_roles
from app.models.user import UserInDB
from app.schemas.analytics import AnalyticsOverview, FarmerAttentionItem, FeedbackReliabilityStats
from app.schemas.response import APIResponse
from app.services.analytics_service import AnalyticsService
from app.utils.responses import success_response

router = APIRouter()


@router.get("/overview", response_model=APIResponse[AnalyticsOverview])
async def overview(
    location: Optional[str] = None,
    crop: Optional[str] = None,
    farm_size_min: Optional[float] = Query(default=None, ge=0),
    farm_size_max: Optional[float] = Query(default=None, ge=0),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: AsyncIOMotorDatabase = Depends(get_db),
    user: UserInDB = Depends(require_roles(["extension_officer", "admin"])),
) -> APIResponse[AnalyticsOverview]:
    service = AnalyticsService(db)
    data = await service.overview(
        location,
        crop,
        farm_size_min,
        farm_size_max,
        from_date,
        to_date,
        actor=user,
    )
    return success_response(data, message="analytics overview")


@router.get("/farmers-needing-attention", response_model=APIResponse[list[FarmerAttentionItem]])
async def farmers_needing_attention(
    location: Optional[str] = None,
    consent_safe: bool = Query(default=True),
    limit: int = Query(default=20, ge=1, le=200),
    db: AsyncIOMotorDatabase = Depends(get_db),
    user: UserInDB = Depends(require_roles(["extension_officer", "admin"])),
) -> APIResponse[list[FarmerAttentionItem]]:
    service = AnalyticsService(db)
    data = await service.farmers_needing_attention(
        location=location,
        consent_safe=consent_safe,
        limit=limit,
        actor=user,
    )
    return success_response(data, message="risk ranking generated")


@router.get("/feedback-reliability", response_model=APIResponse[FeedbackReliabilityStats])
async def feedback_reliability(
    location: Optional[str] = None,
    db: AsyncIOMotorDatabase = Depends(get_db),
    user: UserInDB = Depends(require_roles(["extension_officer", "admin"])),
) -> APIResponse[FeedbackReliabilityStats]:
    service = AnalyticsService(db)
    data = await service.feedback_reliability(location=location, actor=user)
    return success_response(data, message="feedback reliability generated")


def _build_excel(overview: AnalyticsOverview) -> BytesIO:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise HTTPException(status_code=503, detail="Excel export dependency not installed") from exc
    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"
    ws.append(["Metric", "Value"])
    ws.append(["Total Farmers", overview.total_farmers])
    ws.append(["Total Feedback", overview.total_feedback])
    ws.append(["Avg Sustainability", overview.average_sustainability])
    ws.append(["Avg Yield (kg/acre)", overview.average_yield_kg_per_acre])
    ws.append(["Avg Water Usage (L/acre)", overview.average_water_usage_l_per_acre])
    ws.append(["Avg Fertilizer (kg/acre)", overview.average_fertilizer_kg_per_acre])
    ws.append(["At Risk Farmers", overview.at_risk_farmers])

    ws2 = wb.create_sheet("Top Crops")
    ws2.append(["Crop", "Count", "Percentage"])
    for item in overview.top_crops:
        ws2.append([item.crop, item.count, item.percentage])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _build_pdf(overview: AnalyticsOverview) -> BytesIO:
    try:
        from fpdf import FPDF
    except ImportError as exc:
        raise HTTPException(status_code=503, detail="PDF export dependency not installed") from exc
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_font("Helvetica", size=14)
    pdf.cell(0, 10, "KrishiMitra-AI - Analytics Overview", ln=True)
    pdf.set_font("Helvetica", size=11)
    pdf.cell(0, 8, f"Generated: {overview.generated_at.isoformat()}", ln=True)
    pdf.ln(4)
    rows = [
        ("Total Farmers", overview.total_farmers),
        ("Total Feedback", overview.total_feedback),
        ("Avg Sustainability", overview.average_sustainability),
        ("Avg Yield (kg/acre)", overview.average_yield_kg_per_acre),
        ("Avg Water Usage (L/acre)", overview.average_water_usage_l_per_acre),
        ("Avg Fertilizer (kg/acre)", overview.average_fertilizer_kg_per_acre),
        ("At Risk Farmers", overview.at_risk_farmers),
    ]
    for label, value in rows:
        pdf.cell(0, 8, f"{label}: {value}", ln=True)
    pdf.ln(4)
    pdf.set_font("Helvetica", size=12)
    pdf.cell(0, 8, "Top Crops", ln=True)
    pdf.set_font("Helvetica", size=11)
    for item in overview.top_crops:
        pdf.cell(0, 7, f"{item.crop}: {item.count} ({item.percentage}%)", ln=True)

    pdf_bytes = pdf.output(dest="S").encode("latin1")
    output = BytesIO(pdf_bytes)
    output.seek(0)
    return output


@router.get("/export")
async def export_report(
    format: str = Query(default="xlsx", pattern="^(xlsx|pdf)$"),
    location: Optional[str] = None,
    crop: Optional[str] = None,
    farm_size_min: Optional[float] = Query(default=None, ge=0),
    farm_size_max: Optional[float] = Query(default=None, ge=0),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: AsyncIOMotorDatabase = Depends(get_db),
    user: UserInDB = Depends(require_roles(["extension_officer", "admin"])),
) -> Response:
    service = AnalyticsService(db)
    overview = await service.overview(
        location,
        crop,
        farm_size_min,
        farm_size_max,
        from_date,
        to_date,
        actor=user,
    )

    if format == "pdf":
        pdf_stream = _build_pdf(overview)
        return StreamingResponse(
            pdf_stream,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=krishimitra-analytics.pdf"},
        )

    excel_stream = _build_excel(overview)
    return StreamingResponse(
        excel_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=krishimitra-analytics.xlsx"},
    )

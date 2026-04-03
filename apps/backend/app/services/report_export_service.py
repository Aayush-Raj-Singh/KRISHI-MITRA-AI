from __future__ import annotations

import re
from datetime import datetime, timezone
from io import BytesIO
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.core.database import Database
from app.core.logging import get_logger
from app.models.user import UserInDB
from app.schemas.analytics import AnalyticsReportFormat
from app.schemas.dashboard import RegionalInsightsResponse
from app.services.analytics_service import AnalyticsService

logger = get_logger(__name__)


class ReportExportService:
    def __init__(self, db: Database) -> None:
        self._analytics = AnalyticsService(db)

    async def export_regional_insights(
        self,
        *,
        report_format: AnalyticsReportFormat,
        location: Optional[str],
        crop: Optional[str],
        farm_size_min: Optional[float],
        farm_size_max: Optional[float],
        from_date: Optional[str],
        to_date: Optional[str],
        consent_safe: bool,
        limit: int,
        actor: UserInDB | None = None,
    ) -> tuple[bytes, str, str]:
        report = await self._analytics.regional_insights(
            location=location,
            crop=crop,
            farm_size_min=farm_size_min,
            farm_size_max=farm_size_max,
            from_date=from_date,
            to_date=to_date,
            consent_safe=consent_safe,
            limit=limit,
            actor=actor,
        )

        if report_format == AnalyticsReportFormat.pdf:
            content = self._build_pdf(report)
            media_type = "application/pdf"
        else:
            content = self._build_workbook(report)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        filename = self._build_filename(report_format, location=location, crop=crop)
        logger.info(
            "analytics_report_exported",
            format=report_format.value,
            filename=filename,
            location=location,
            crop=crop,
        )
        return content, media_type, filename

    def _build_filename(
        self,
        report_format: AnalyticsReportFormat,
        *,
        location: Optional[str],
        crop: Optional[str],
    ) -> str:
        segments = ["regional-insights"]
        if location:
            segments.append(self._slug(location))
        if crop:
            segments.append(self._slug(crop))
        segments.append(self._slug_date())
        return f"{'-'.join(filter(None, segments))}.{report_format.value}"

    @staticmethod
    def _slug(value: str) -> str:
        normalized = re.sub(r"[^a-z0-9]+", "-", value.strip().lower())
        return normalized.strip("-") or "all"

    @staticmethod
    def _slug_date() -> str:
        return datetime.now(timezone.utc).strftime("%Y%m%d")

    @staticmethod
    def _format_filters(report: RegionalInsightsResponse) -> str:
        filters = report.overview.filters or {}
        visible = []
        for key, value in filters.items():
            if value in (None, "", []):
                continue
            label = key.replace("_", " ").title()
            visible.append(f"{label}: {value}")
        if not visible:
            return "Filters: default regional scope"
        return "Filters: " + "; ".join(visible)

    @staticmethod
    def _metric_rows(report: RegionalInsightsResponse) -> list[list[str]]:
        overview = report.overview
        return [
            ["Metric", "Value"],
            ["Total farmers", str(overview.total_farmers)],
            ["Total feedback", str(overview.total_feedback)],
            ["Average sustainability", f"{overview.average_sustainability:.2f}"],
            ["Average yield (kg/acre)", f"{overview.average_yield_kg_per_acre:.2f}"],
            ["Average water use (L/acre)", f"{overview.average_water_usage_l_per_acre:.2f}"],
            ["Average fertilizer (kg/acre)", f"{overview.average_fertilizer_kg_per_acre:.2f}"],
            ["At-risk farmers", str(overview.at_risk_farmers)],
        ]

    @staticmethod
    def _top_crop_rows(report: RegionalInsightsResponse) -> list[list[str]]:
        rows = [["Crop", "Recommendations", "Share %"]]
        if not report.overview.top_crops:
            rows.append(["No crop adoption trend data available", "-", "-"])
            return rows
        for item in report.overview.top_crops:
            rows.append([item.crop, str(item.count), f"{item.percentage:.2f}"])
        return rows

    @staticmethod
    def _attention_rows(report: RegionalInsightsResponse) -> list[list[str]]:
        rows = [["Farmer", "Location", "Risk", "Sustainability", "Yield trend %", "Reasons"]]
        if not report.farmers_needing_attention:
            rows.append(["No farmers currently require manual attention", "-", "-", "-", "-", "-"])
            return rows
        for farmer in report.farmers_needing_attention:
            rows.append(
                [
                    farmer.name,
                    farmer.location,
                    f"{farmer.risk_score:.2f}",
                    f"{farmer.sustainability_score:.2f}",
                    f"{farmer.yield_trend_percent:.2f}",
                    ", ".join(farmer.reasons),
                ]
            )
        return rows

    @staticmethod
    def _reliability_rows(report: RegionalInsightsResponse) -> list[list[str]]:
        reliability = report.feedback_reliability
        distribution = " | ".join(
            f"{star}*={count}"
            for star, count in sorted(
                reliability.rating_distribution.items(), key=lambda item: item[0]
            )
        )
        return [
            ["Metric", "Value"],
            ["Total feedback", str(reliability.total_feedback)],
            ["Average rating", f"{reliability.average_rating:.2f}"],
            ["Negative outcome rate %", f"{reliability.negative_outcome_rate:.2f}"],
            ["Pending expert reviews", str(reliability.expert_review_pending)],
            ["Rating distribution", distribution or "No ratings yet"],
        ]

    def _build_workbook(self, report: RegionalInsightsResponse) -> bytes:
        workbook = Workbook()
        overview_sheet = workbook.active
        overview_sheet.title = "Overview"
        self._populate_sheet(
            overview_sheet,
            title="Regional Insights Overview",
            subtitle=self._format_filters(report),
            rows=self._metric_rows(report),
        )

        top_crops_sheet = workbook.create_sheet("Top Crops")
        self._populate_sheet(
            top_crops_sheet,
            title="Crop Adoption Trends",
            subtitle=f"Generated at {report.generated_at.isoformat()}",
            rows=self._top_crop_rows(report),
        )

        attention_sheet = workbook.create_sheet("Farmer Attention")
        self._populate_sheet(
            attention_sheet,
            title="Farmers Needing Attention",
            subtitle="Masked identities remain protected unless explicit farmer consent is stored.",
            rows=self._attention_rows(report),
        )

        reliability_sheet = workbook.create_sheet("Reliability")
        self._populate_sheet(
            reliability_sheet,
            title="Feedback Reliability",
            subtitle="Aggregate reliability and manual review signals.",
            rows=self._reliability_rows(report),
        )

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _populate_sheet(
        self, sheet, *, title: str, subtitle: str, rows: Iterable[Iterable[str]]
    ) -> None:
        header_fill = PatternFill(fill_type="solid", fgColor="1B6B3A")
        title_font = Font(size=14, bold=True)
        header_font = Font(bold=True, color="FFFFFF")

        sheet["A1"] = title
        sheet["A1"].font = title_font
        sheet["A2"] = subtitle
        sheet["A2"].alignment = Alignment(wrap_text=True)

        row_start = 4
        for row_index, row in enumerate(rows, start=row_start):
            values = list(row)
            for col_index, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_index, column=col_index, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if row_index == row_start:
                    cell.fill = header_fill
                    cell.font = header_font
            if row_index == row_start:
                sheet.freeze_panes = f"A{row_index + 1}"

        self._autosize_columns(sheet)

    @staticmethod
    def _autosize_columns(sheet) -> None:
        for column_cells in sheet.columns:
            column_letter = column_cells[0].column_letter
            max_length = 0
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 42)

    def _build_pdf(self, report: RegionalInsightsResponse) -> bytes:
        buffer = BytesIO()
        document = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=24,
            rightMargin=24,
            topMargin=24,
            bottomMargin=24,
        )
        styles = getSampleStyleSheet()

        elements = [
            Paragraph("KrishiMitra AI Regional Insights Report", styles["Title"]),
            Spacer(1, 8),
            Paragraph(self._format_filters(report), styles["BodyText"]),
            Paragraph(f"Generated at: {report.generated_at.isoformat()}", styles["BodyText"]),
            Paragraph(
                "<i>Privacy note: this report only exposes individual identifiers where consent exists.</i>",
                styles["BodyText"],
            ),
            Spacer(1, 12),
            Paragraph("Overview", styles["Heading2"]),
            self._build_table(self._metric_rows(report), col_widths=[190, 120]),
            Spacer(1, 12),
            Paragraph("Crop Adoption Trends", styles["Heading2"]),
            self._build_table(self._top_crop_rows(report), col_widths=[220, 120, 100]),
            Spacer(1, 12),
            Paragraph("Feedback Reliability", styles["Heading2"]),
            self._build_table(self._reliability_rows(report), col_widths=[220, 340]),
            Spacer(1, 12),
            Paragraph("Farmers Needing Attention", styles["Heading2"]),
            self._build_table(
                self._attention_rows(report),
                col_widths=[100, 110, 60, 78, 80, 300],
            ),
        ]

        document.build(elements)
        return buffer.getvalue()

    @staticmethod
    def _build_table(rows: list[list[str]], *, col_widths: list[int]) -> Table:
        table = Table(rows, repeatRows=1, colWidths=col_widths)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B6B3A")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D4D9D6")),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#F5F7F6")],
                    ),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        return table
